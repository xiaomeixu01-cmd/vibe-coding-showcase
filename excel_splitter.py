#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel文件拆分脚本
功能：将Excel文件按每1000行拆分为多个小文件
"""

import os
import sys
from openpyxl import load_workbook
from openpyxl import Workbook

def split_excel_file(file_path, rows_per_file=1000):
    """
    拆分Excel文件
    
    Args:
        file_path: Excel文件路径
        rows_per_file: 每个文件的最大行数，默认1000
    """
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            print(f"错误：文件 {file_path} 不存在")
            return
        
        # 检查文件扩展名是否为.xlsx
        if not file_path.endswith('.xlsx'):
            print("错误：仅支持.xlsx格式的Excel文件")
            return
        
        # 获取文件所在目录
        file_dir = os.path.dirname(file_path)
        if file_dir == '':
            file_dir = '.'
        
        # 获取文件名（不含扩展名）
        file_name = os.path.basename(file_path)
        base_name = os.path.splitext(file_name)[0]
        
        print(f"正在处理文件：{file_name}")
        print(f"拆分后每个文件最多包含 {rows_per_file} 行数据")
        
        # 加载Excel文件
        wb = load_workbook(file_path)
        # 获取第一个工作表
        ws = wb.active
        
        # 获取总行数
        total_rows = ws.max_row
        print(f"文件共有 {total_rows} 行数据")
        
        # 检查是否有足够的数据需要拆分
        if total_rows <= rows_per_file:
            print("提示：文件行数不足，无需拆分")
            return
        
        # 获取表头（第一行）
        header = []
        for cell in ws[1]:
            header.append(cell.value)
        
        # 计算需要拆分的文件数
        file_count = (total_rows - 1) // rows_per_file + 1  # 减1是因为表头不计入数据行
        print(f"将拆分为 {file_count} 个文件")
        
        # 开始拆分
        for i in range(file_count):
            # 计算当前文件的起始行和结束行
            start_row = i * rows_per_file + 2  # +2是因为数据从第二行开始，且Python是从1开始计数
            end_row = min((i + 1) * rows_per_file + 1, total_rows)  # +1是因为包含表头
            
            # 创建新的工作簿
            new_wb = Workbook()
            new_ws = new_wb.active
            
            # 写入表头
            new_ws.append(header)
            
            # 写入数据
            for row in range(start_row, end_row + 1):
                row_data = []
                for cell in ws[row]:
                    row_data.append(cell.value)
                new_ws.append(row_data)
            
            # 生成新文件名
            new_file_name = f"{base_name}-{i+1}.xlsx"
            new_file_path = os.path.join(file_dir, new_file_name)
            
            # 保存新文件
            new_wb.save(new_file_path)
            print(f"已生成文件：{new_file_name}")
        
        print("\n拆分完成！")
        
    except Exception as e:
        print(f"错误：{str(e)}")

def main():
    """
    主函数
    """
    # 这里可以修改为你要拆分的Excel文件路径
    # 例如：excel_file = r"C:\Users\用户名\Desktop\data.xlsx"
    excel_file = r"C:\Users\Administrator\Desktop\表格拆分\商户号.xlsx"
    
    # 调用拆分函数
    split_excel_file(excel_file)

if __name__ == "__main__":
    main()
